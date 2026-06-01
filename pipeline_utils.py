from __future__ import annotations

import colorsys
import hashlib
import os
import re
from pathlib import Path

import pandas as pd
import pymupdf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


STOPWORDS_PATH = Path(__file__).resolve().parent / "english_stopwords.txt"

with STOPWORDS_PATH.open("r", encoding="utf-8") as handle:
    STOP_WORDS = {line.strip() for line in handle if line.strip()}


def split_abbreviation_v2(abbr: str) -> list[str]:
    components = []
    i = 0

    if abbr.endswith("s") and len(abbr) > 1:
        abbr = abbr[:-1]

    abbr = abbr.replace("-", "")

    if len(abbr) == 3 and abbr[0].isupper() and abbr[1].islower() and abbr[2].isupper():
        return [abbr[0], abbr[1], abbr[2]]

    while i < len(abbr):
        if i > 0 and i < len(abbr) - 1 and abbr[i].isdigit() and abbr[i - 1].isalpha() and abbr[i + 1].isalpha():
            components.append(abbr[i])
            i += 1
            continue

        if i < len(abbr) - 1 and abbr[i].lower() == "v" and abbr[i + 1].isdigit() and i + 2 >= len(abbr):
            if components and components[-1].isupper():
                last_component = components.pop()
                components.extend(list(last_component))

            j = i + 1
            while j < len(abbr) and abbr[j].isdigit():
                j += 1
            components.append(f"v{abbr[i + 1:j]}")
            i = j
            continue

        if i + 1 < len(abbr) and abbr[i : i + 2].lower() == abbr[i : i + 2].lower() * 2:
            components.append(abbr[i : i + 2])
            i += 2
            continue

        if abbr[i].isupper():
            if i + 1 < len(abbr) and abbr[i + 1].islower():
                j = i + 2
                while j < len(abbr) and abbr[j].islower():
                    j += 1
                components.append(abbr[i:j])
                i = j
            else:
                j = i + 1
                while j < len(abbr) and abbr[j].isupper():
                    j += 1
                if j - i > 1:
                    components.extend(list(abbr[i:j]))
                else:
                    components.append(abbr[i])
                i = j
        else:
            components.append(abbr[i])
            i += 1

    return components


def generate_combinations(components: list[str]) -> list[list[str]]:
    result = []

    def merge_components(start: int, end: int) -> str:
        return "".join(components[start:end])

    for i in range(len(components)):
        for j in range(i + 2, len(components) + 1):
            result.append(components[:i] + [merge_components(i, j)] + components[j:])

    return result


def find_full_form_v2(context: str, components: list[str]) -> str | None:
    context = re.sub(r"[^\w\s\-]", " ", context)
    context = " ".join(context.split())

    words = []
    for word in context.split():
        if "-" in word:
            words.extend(word.split("-"))
        else:
            words.append(word)
    words = [word for word in words if word.strip()]

    if not words:
        return None

    def matches_component(word: str, component: str, include_stopwords: bool = False, case_sensitive: bool = True) -> bool:
        if not word or not component:
            return False

        if component.lower() == "o" and word.lower() == "of":
            return True
        if component == "2" and word.lower() == "to":
            return True
        if not include_stopwords and word.lower() in STOP_WORDS:
            return False
        if component.isdigit():
            return component in word or word.lower().startswith("to-" + component)
        if component.startswith("v") and component[1:].isdigit():
            version_num = component[1:]
            return "v" + version_num in word
        if component.upper() == "X":
            lower_word = word.lower()
            return lower_word.startswith(("ex", "xt")) or lower_word == "everything" or lower_word.startswith("every")
        if len(component) == 1 and component.isupper():
            return word == component or word[0].upper() == component
        if case_sensitive:
            return word.startswith(component) or component == word[: len(component)]
        return word.lower().startswith(component.lower()) or component.lower() == word.lower()[: len(component)]

    def find_match_indices(test_words: list[str], test_components: list[str], include_stopwords: bool = False, case_sensitive: bool = True) -> list[int] | None:
        matches = []
        used_indices = set()
        num_components = len(test_components)

        for idx, component in enumerate(test_components):
            found = False
            for word_idx in range(len(test_words) - 1, -1, -1):
                if word_idx in used_indices:
                    continue
                if matches_component(test_words[word_idx], component, include_stopwords, case_sensitive):
                    matches.append(word_idx)
                    used_indices.add(word_idx)
                    found = True
                    break

            if not found:
                if idx == num_components - 1 and matches:
                    return sorted(matches)
                return None

        return sorted(matches)

    def extend_match(start_idx: int, end_idx: int) -> tuple[int, int]:
        if end_idx + 1 < len(words):
            next_words = []
            i = end_idx + 1
            while i < len(words):
                word = words[i]
                if word[0].isupper() or word.lower() not in STOP_WORDS or (i > 0 and "-" in words[i - 1]):
                    next_words.append(word)
                    i += 1
                else:
                    break
            if next_words:
                end_idx += len(next_words)
        return start_idx, end_idx

    match_indices = find_match_indices(words, components, include_stopwords=False, case_sensitive=True)
    if match_indices is None:
        match_indices = find_match_indices(words, components, include_stopwords=True, case_sensitive=True)
    if match_indices is None:
        match_indices = find_match_indices(words, components, include_stopwords=True, case_sensitive=False)
    if match_indices is None:
        for alt_components in generate_combinations(components):
            match_indices = find_match_indices(words, alt_components, include_stopwords=True, case_sensitive=False)
            if match_indices is not None:
                break

    if match_indices:
        start_idx, end_idx = extend_match(min(match_indices), max(match_indices))
        return " ".join(words[start_idx : end_idx + 1])

    return None


def string_to_light_color_hex(text: str) -> str:
    digest = hashlib.md5(text.encode("utf-8")).hexdigest()
    hue = int(digest[0:2], 16) / 255
    sat = int(digest[2:4], 16) / 255
    lit = int(digest[4:6], 16) / 255
    red, green, blue = colorsys.hls_to_rgb(hue, 0.75 + (lit * 0.1), 0.4 + (sat * 0.2))
    return "#{:02X}{:02X}{:02X}".format(int(red * 255), int(green * 255), int(blue * 255))


def hex_to_rgb01(hex_color: str) -> tuple[float, float, float]:
    hex_color = hex_color.lstrip("#")
    return (
        int(hex_color[0:2], 16) / 255,
        int(hex_color[2:4], 16) / 255,
        int(hex_color[4:6], 16) / 255,
    )


def process_pdf(pdf_path: str | Path, output_folder: str | Path, max_pages: int = 100) -> dict[str, dict[str, object]]:
    output_folder = Path(output_folder)
    page_text_dir = output_folder / "page_texts"
    page_text_dir.mkdir(parents=True, exist_ok=True)

    collection: dict[str, dict[str, object]] = {}
    pattern_extract_abbr = r"(.*?)\(([A-Z][A-Za-z0-9\-\+\/\.>=<]*?)\)"

    doc = pymupdf.open(pdf_path)
    try:
        if len(doc) > max_pages:
            raise ValueError(f"PDF has {len(doc)} pages. Please upload a PDF with {max_pages} pages or fewer.")

        for page_index in range(len(doc)):
            text = doc[page_index].get_text()
            clean_text = " ".join(text.split())
            (page_text_dir / f"page_{page_index + 1}.txt").write_text(clean_text, encoding="utf-8")

            for match in re.finditer(pattern_extract_abbr, clean_text):
                context_before = match.group(1).strip()[-250:]
                abbr = match.group(2)
                components = split_abbreviation_v2(abbr)
                full_form = find_full_form_v2(context_before, components)
                key = abbr.upper()

                if key not in collection:
                    collection[key] = {
                        "Abbreviation": abbr,
                        "Full Form": full_form,
                        "Pages": [],
                        "Color": string_to_light_color_hex(f"{abbr}_{full_form}"),
                    }

                if full_form and not collection[key]["Full Form"]:
                    collection[key]["Full Form"] = full_form
                    collection[key]["Color"] = string_to_light_color_hex(f"{abbr}_{full_form}")

                collection[key]["Pages"].append(page_index + 1)
    finally:
        doc.close()

    return collection


def export_to_excel(collection: dict[str, dict[str, object]], output_excel: str | Path) -> None:
    rows = []
    for item in collection.values():
        rows.append(
            {
                "Abbreviation": item["Abbreviation"],
                "Full Form": item["Full Form"] or "",
                "Pages": ", ".join(map(str, item["Pages"])),
                "Color": item["Color"],
            }
        )

    if not rows:
        rows.append({"Abbreviation": "No abbreviations found", "Full Form": "", "Pages": "", "Color": ""})

    df = pd.DataFrame(rows)
    df.to_excel(output_excel, index=False)

    workbook = load_workbook(output_excel)
    worksheet = workbook.active
    color_col = None
    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == "Color":
            color_col = idx
            break

    if color_col:
        for row in range(2, worksheet.max_row + 1):
            hex_color = worksheet.cell(row=row, column=color_col).value
            if hex_color:
                fill = PatternFill(start_color=hex_color.replace("#", ""), end_color=hex_color.replace("#", ""), fill_type="solid")
                worksheet.cell(row=row, column=color_col).fill = fill

    workbook.save(output_excel)


def highlight_terms_in_pdf(input_pdf: str | Path, output_pdf: str | Path, items: list[dict[str, str]]) -> None:
    doc = pymupdf.open(input_pdf)
    try:
        for page in doc:
            for item in items:
                full_form = item.get("full_form", "")
                if not full_form.strip():
                    continue

                rgb = hex_to_rgb01(item["color"])
                full_quads = page.search_for(full_form, quads=True, flags=0)
                if full_quads:
                    annot = page.add_highlight_annot(full_quads)
                    annot.set_colors(stroke=rgb)
                    annot.update()

        doc.save(output_pdf)
    finally:
        doc.close()


def collection_to_highlight_items(collection: dict[str, dict[str, object]]) -> list[dict[str, str]]:
    return [
        {
            "abbr": str(item["Abbreviation"]),
            "full_form": str(item["Full Form"] or ""),
            "color": str(item["Color"]),
        }
        for item in collection.values()
    ]
